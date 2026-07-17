# attendance/views.py

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404 # type: ignore
from django.utils import timezone # type: ignore
from datetime import timedelta, datetime
from attendance.models import Attendance, AttendanceAttempt, AttendanceReport, GeolocationLog
from attendance.serializers import (
    AttendanceSerializer,
    AttendanceAttemptSerializer,
    AttendanceReportSerializer,
    GeolocationLogSerializer
)
from courses.models import Lecture, Course, Enrollment
from notifications.models import Notification
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.styles import getSampleStyleSheet # type: ignore
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer # type: ignore
from reportlab.lib.units import inch # type: ignore
import openpyxl # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment # type: ignore
from io import BytesIO
import math
import base64
import json

class AttendanceViewSet(viewsets.ModelViewSet):
    """Attendance management viewset"""
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]
    MAX_LOCATION_ACCURACY_METERS = 50

    def get_queryset(self):
        """Filter attendance records"""
        user = self.request.user
        if user.role == 'student':
            return Attendance.objects.filter(student=user)
        elif user.role == 'lecturer':
            return Attendance.objects.filter(
                lecture__course__lecturer=user
            )
        return Attendance.objects.all()

    @action(detail=False, methods=['post'])
    def mark_attendance(self, request):
        """Mark student attendance with QR code"""
        if request.user.role != 'student':
            return Response(
                {'error': 'Only students can mark attendance'},
                status=status.HTTP_403_FORBIDDEN
            )

        lecture_id = request.data.get('lecture_id')
        qr_code_data = request.data.get('qr_code_data')
        latitude = request.data.get('latitude')
        longitude = request.data.get('longitude')
        accuracy = request.data.get('accuracy')

        if not qr_code_data:
            return self._record_failed_attempt(
                request=request,
                lecture=None,
                reason='invalid_qr',
                message='QR code data is required'
            )

        if not latitude or not longitude:
            return self._record_failed_attempt(
                request=request,
                lecture=None,
                reason='missing_location',
                message='Location is required. Please allow location access and retry.'
            )

        decoded_qr = self._decode_qr_code(qr_code_data)
        if not decoded_qr:
            return self._record_failed_attempt(
                request=request,
                lecture=None,
                reason='invalid_qr',
                message='Invalid QR code data',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        qr_lecture_id = decoded_qr.get('lecture_id')
        if lecture_id and int(lecture_id) != int(qr_lecture_id):
            return self._record_failed_attempt(
                request=request,
                lecture=None,
                reason='invalid_qr',
                message='QR code does not match the selected lecture',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        lecture_id = qr_lecture_id

        # Get lecture
        lecture = get_object_or_404(Lecture, lecture_id=lecture_id)

        if not lecture.course.enrollments.filter(student=request.user, status='active').exists():
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='unauthorized',
                message='You are not enrolled in this course.',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy,
                allowed_radius=lecture.allowed_radius
            )

        # Verify QR code
        if lecture.qr_status == 'closed':
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='closed_qr',
                message='This attendance session has ended',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        if not lecture.qr_is_valid or lecture.qr_code_data != qr_code_data:
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='expired_qr',
                message='QR code is invalid or has expired',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        if not lecture.venue_latitude or not lecture.venue_longitude:
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='lecturer_location_missing',
                message='Lecturer location was not captured for this QR session',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        # Check for duplicate attendance
        existing = Attendance.objects.filter(
            lecture=lecture,
            student=request.user
        ).first()

        if existing:
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='already_marked',
                message='You already marked attendance for this lecture.',
                status_value='duplicate',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy,
                allowed_radius=lecture.allowed_radius
            )

        try:
            accuracy_value = float(accuracy) if accuracy not in [None, ''] else None
        except (TypeError, ValueError):
            accuracy_value = None

        if accuracy_value is None or accuracy_value > self.MAX_LOCATION_ACCURACY_METERS:
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='poor_accuracy',
                message=(
                    f'Your GPS accuracy is too low. Please retry with accuracy within '
                    f'{self.MAX_LOCATION_ACCURACY_METERS}m.'
                ),
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy,
                allowed_radius=lecture.allowed_radius
            )

        # Verify location if geolocation data provided
        location_verified = False
        distance_from_venue = None

        try:
            distance_from_venue = self._calculate_distance(
                float(latitude), float(longitude),
                float(lecture.venue_latitude), float(lecture.venue_longitude)
            )
        except (TypeError, ValueError):
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='invalid_location',
                message='Invalid student location. Please retry with location enabled.',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy
            )

        if distance_from_venue <= lecture.allowed_radius:
            location_verified = True
        else:
            GeolocationLog.objects.create(
                attendance=None,
                student=request.user,
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy,
                distance_from_venue=distance_from_venue,
                verification_status='failed',
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            return self._record_failed_attempt(
                request=request,
                lecture=lecture,
                reason='outside_radius',
                message='You are outside the allowed attendance radius. Please move closer and retry.',
                latitude=latitude,
                longitude=longitude,
                accuracy=accuracy,
                distance_from_venue=distance_from_venue,
                allowed_radius=lecture.allowed_radius
            )

        late_cutoff = lecture.qr_generated_at + timedelta(minutes=30) if lecture.qr_generated_at else None
        is_late = bool(late_cutoff and timezone.now() > late_cutoff)

        # Create attendance record
        attendance = Attendance.objects.create(
            lecture=lecture,
            student=request.user,
            latitude=latitude,
            longitude=longitude,
            location_verified=location_verified,
            distance_from_venue=distance_from_venue,
            ip_address=self._get_client_ip(request),
            is_late=is_late
        )

        # Log geolocation
        GeolocationLog.objects.create(
            attendance=attendance,
            student=request.user,
            latitude=latitude,
            longitude=longitude,
            accuracy=accuracy,
            distance_from_venue=distance_from_venue,
            verification_status='verified' if location_verified else 'failed',
            ip_address=self._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        AttendanceAttempt.objects.create(
            lecture=lecture,
            student=request.user,
            attendance=attendance,
            status='success',
            reason='marked',
            message='Attendance marked successfully',
            latitude=latitude,
            longitude=longitude,
            accuracy=accuracy,
            distance_from_venue=distance_from_venue,
            allowed_radius=lecture.allowed_radius,
            ip_address=self._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        return Response(
            {
                'success': True,
                'message': 'Attendance marked successfully',
                'distance': distance_from_venue,
                'allowed_radius': lecture.allowed_radius,
                'attendance': AttendanceSerializer(attendance).data
            },
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['get'])
    def my_scan_history(self, request):
        if request.user.role != 'student':
            return Response(
                {'error': 'Only students can view scan history'},
                status=status.HTTP_403_FORBIDDEN
            )
        attempts = AttendanceAttempt.objects.filter(student=request.user)[:50]
        return Response(AttendanceAttemptSerializer(attempts, many=True).data)

    def _record_failed_attempt(
        self,
        request,
        lecture,
        reason,
        message,
        status_value='failed',
        latitude=None,
        longitude=None,
        accuracy=None,
        distance_from_venue=None,
        allowed_radius=None
    ):
        attempt = AttendanceAttempt.objects.create(
            lecture=lecture,
            student=request.user,
            status=status_value,
            reason=reason,
            message=message,
            latitude=latitude,
            longitude=longitude,
            accuracy=accuracy,
            distance_from_venue=distance_from_venue,
            allowed_radius=allowed_radius,
            ip_address=self._get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )

        Notification.objects.create(
            recipient=request.user,
            sender=lecture.course.lecturer if lecture else None,
            title='Attendance scan rejected',
            message=message,
            notification_type='attendance',
            priority='high' if reason in ['outside_radius', 'poor_accuracy'] else 'normal',
            related_object_type='attendance_attempt',
            related_object_id=attempt.attempt_id
        )

        return Response(
            {
                'error': message,
                'code': reason,
                'attempt_id': attempt.attempt_id,
                'distance': distance_from_venue,
                'allowed_radius': allowed_radius
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    @staticmethod
    def _decode_qr_code(qr_code_data):
        try:
            decoded = base64.b64decode(qr_code_data).decode('utf-8')
            data = json.loads(decoded)
            if data.get('type') != 'attendance' or not data.get('lecture_id'):
                return None
            return data
        except Exception:
            return None

    @action(detail=False, methods=['get'])
    def course_attendance(self, request):
        """Get attendance records for a specific course"""
        course_id = request.query_params.get('course_id')
        if not course_id:
            return Response(
                {'error': 'course_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        course = get_object_or_404(Course, course_id=course_id)

        # Check permission
        if request.user.role == 'lecturer' and course.lecturer != request.user:
            return Response(
                {'error': 'Not authorized'},
                status=status.HTTP_403_FORBIDDEN
            )

        lectures = Lecture.objects.filter(course=course).values_list(
            'lecture_id', flat=True
        )
        attendances = Attendance.objects.filter(lecture_id__in=lectures)

        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def my_course_summary(self, request):
        """Return a student's attendance percentage per enrolled course."""
        if request.user.role != 'student':
            return Response(
                {'error': 'Only students can view their attendance summary'},
                status=status.HTTP_403_FORBIDDEN
            )

        summaries = []
        enrollments = Enrollment.objects.filter(
            student=request.user,
            status='active'
        ).select_related('course', 'course__lecturer')

        for enrollment in enrollments:
            course = enrollment.course
            lecture_ids = list(
                course.lectures.filter(lecture_date__lte=timezone.localdate())
                .values_list('lecture_id', flat=True)
            )
            total = len(lecture_ids)
            attendance_qs = Attendance.objects.filter(
                lecture_id__in=lecture_ids,
                student=request.user,
                location_verified=True
            )
            attended = attendance_qs.count()
            late = attendance_qs.filter(is_late=True).count()
            missed = max(total - attended, 0)
            percentage = round((attended / total) * 100) if total else 0

            summaries.append({
                'course': {
                    'course_id': course.course_id,
                    'course_code': course.course_code,
                    'course_name': course.course_name,
                    'department': course.department,
                    'semester': course.semester,
                    'lecturer_name': course.lecturer.full_name,
                },
                'total_lectures': total,
                'attended': attended,
                'missed': missed,
                'late': late,
                'percentage': percentage
            })

        return Response(summaries)

    @action(detail=False, methods=['get'])
    def missed_lectures(self, request):
        """Return lectures a student missed in enrolled courses."""
        if request.user.role != 'student':
            return Response(
                {'error': 'Only students can view missed lectures'},
                status=status.HTTP_403_FORBIDDEN
            )

        enrolled_course_ids = Enrollment.objects.filter(
            student=request.user,
            status='active'
        ).values_list('course_id', flat=True)
        attended_lecture_ids = Attendance.objects.filter(
            student=request.user,
            location_verified=True
        ).values_list('lecture_id', flat=True)
        missed = Lecture.objects.filter(
            course_id__in=enrolled_course_ids,
            lecture_date__lte=timezone.localdate()
        ).exclude(lecture_id__in=attended_lecture_ids).order_by('-lecture_date', 'start_time')[:30]

        return Response([{
            'lecture_id': lecture.lecture_id,
            'topic': lecture.topic,
            'lecture_date': lecture.lecture_date,
            'start_time': lecture.start_time,
            'course_code': lecture.course.course_code,
            'course_name': lecture.course.course_name,
            'location': lecture.location
        } for lecture in missed])

    @action(detail=False, methods=['post'])
    def generate_report(self, request):
        """Generate attendance report"""
        report_type = request.data.get('report_type')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        report_format = request.data.get('format', 'pdf')
        filters = {
            'course_id': request.data.get('course_id'),
            'lecture_id': request.data.get('lecture_id'),
            'student_id': request.data.get('student_id'),
            'department': request.data.get('department')
        }

        if not all([report_type, start_date, end_date]):
            return Response(
                {'error': 'Missing required fields'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Generate report file
        if report_format == 'pdf':
            file_obj = self._generate_pdf_report(
                report_type, start_date, end_date, request.user, filters
            )
            filename = f"attendance_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        else:
            file_obj = self._generate_excel_report(
                report_type, start_date, end_date, request.user, filters
            )
            filename = f"attendance_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Save report record
        report = AttendanceReport.objects.create(
            report_type=report_type,
            generated_by=request.user,
            report_title=f"{report_type.title()} Attendance Report",
            filters={'start_date': start_date, 'end_date': end_date, **filters},
            file_format=report_format
        )

        return Response(
            {
                'success': True,
                'report_id': report.report_id,
                'filename': filename,
                'file': file_obj.getvalue().hex() if hasattr(file_obj, 'getvalue') else None
            },
            status=status.HTTP_201_CREATED
        )

    @staticmethod
    def _calculate_distance(lat1, lon1, lat2, lon2):
        """Calculate distance between two coordinates using Haversine formula"""
        R = 6371000  # Earth radius in meters
        
        phi1 = math.radians(lat1)
        phi2 = math.radians(lat2)
        delta_phi = math.radians(lat2 - lat1)
        delta_lambda = math.radians(lon2 - lon1)

        a = math.sin(delta_phi / 2) ** 2 + \
            math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

        return round(R * c)  # Return distance in meters

    @staticmethod
    def _get_client_ip(request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    @staticmethod
    def _filtered_attendances(start_date, end_date, user, filters=None):
        filters = filters or {}
        attendances = Attendance.objects.filter(
            timestamp__date__range=[start_date, end_date]
        ).select_related('student', 'lecture', 'lecture__course')

        if user.role == 'student':
            attendances = attendances.filter(student=user)
        elif user.role == 'lecturer':
            attendances = attendances.filter(lecture__course__lecturer=user)

        if filters.get('course_id'):
            attendances = attendances.filter(lecture__course_id=filters['course_id'])
        if filters.get('lecture_id'):
            attendances = attendances.filter(lecture_id=filters['lecture_id'])
        if filters.get('student_id'):
            attendances = attendances.filter(student_id=filters['student_id'])
        if filters.get('department'):
            attendances = attendances.filter(lecture__course__department__icontains=filters['department'])

        return attendances

    @staticmethod
    def _generate_pdf_report(report_type, start_date, end_date, user, filters=None):
        """Generate PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("Attendance Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))

        # Data
        data = [['Date', 'Course', 'Student', 'Status', 'Time']]
        attendances = AttendanceViewSet._filtered_attendances(start_date, end_date, user, filters)

        for att in attendances:
            data.append([
                att.timestamp.strftime('%Y-%m-%d'),
                att.lecture.course.course_code,
                att.student.full_name,
                'Present' if not att.is_late else 'Late',
                att.timestamp.strftime('%H:%M:%S')
            ])

        # Table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _generate_excel_report(report_type, start_date, end_date, user, filters=None):
        """Generate Excel report"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance"

        # Headers
        headers = ['Date', 'Course', 'Department', 'Student', 'Status', 'Time', 'Location Verified']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        attendances = AttendanceViewSet._filtered_attendances(start_date, end_date, user, filters)

        for row, att in enumerate(attendances, 2):
            worksheet.cell(row=row, column=1).value = att.timestamp.strftime('%Y-%m-%d')
            worksheet.cell(row=row, column=2).value = att.lecture.course.course_code
            worksheet.cell(row=row, column=3).value = att.lecture.course.department
            worksheet.cell(row=row, column=4).value = att.student.full_name
            worksheet.cell(row=row, column=5).value = 'Present' if not att.is_late else 'Late'
            worksheet.cell(row=row, column=6).value = att.timestamp.strftime('%H:%M:%S')
            worksheet.cell(row=row, column=7).value = 'Yes' if att.location_verified else 'No'

        # Adjust columns
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 22
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 20

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
# Add to AttendanceViewSet in attendance/views.py

    @action(detail=False, methods=['get'])
    def course_analytics(self, request):
        """Get analytics for a specific course"""
        from attendance.analytics import AttendanceAnalytics
        
        course_id = request.query_params.get('course_id')
        if not course_id:
            return Response(
                {'error': 'course_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check permission
        course = get_object_or_404(Course, course_id=course_id)
        if request.user.role == 'lecturer' and course.lecturer != request.user:
            return Response(
                {'error': 'Not authorized'},
                status=status.HTTP_403_FORBIDDEN
            )

        analytics = AttendanceAnalytics.get_course_statistics(course_id)
        return Response(analytics, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def department_analytics(self, request):
        """Get analytics for a department"""
        from attendance.analytics import AttendanceAnalytics
        
        department = request.query_params.get('department')
        if not department:
            return Response(
                {'error': 'department is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Only admin can view department analytics
        if request.user.role != 'admin':
            return Response(
                {'error': 'Not authorized'},
                status=status.HTTP_403_FORBIDDEN
            )

        analytics = AttendanceAnalytics.get_department_statistics(department)
        return Response(analytics, status=status.HTTP_200_OK)
